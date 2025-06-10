

from __future__ import annotations

import hashlib
import io
import json
import re
from typing import Dict, List, Optional

import streamlit as st
from openpyxl import load_workbook


# Regex pattern per criteri
########################################################

RE_SUB = re.compile(r"^\s*CRITERIO\s+([A-Z]\d(?:\.\d+)*)", re.IGNORECASE)
RE_MAIN = re.compile(r"^\s*([A-Z]\d+)\s*[- ]\s*(.+)")


# Helper
##################

def _workbook_from_upload(buf: io.BytesIO):
    return load_workbook(buf, data_only=True, read_only=True)


def _parse_records(sheet, col_letter: str, row_start: int, row_end: int) -> List[Dict[str, str]]:
    records: List[Dict[str, str]] = []
    col_letter = col_letter.upper()
    pending: Optional[str] = None

    for row in range(row_start, row_end + 1):
        value = sheet[f"{col_letter}{row}"].value
        if value is None:
            continue
        text = str(value).strip()
        if not text:
            continue

        if pending:
            records.append({pending: text})
            pending = None
            continue

        m_sub = RE_SUB.match(text)
        if m_sub:
            pending = m_sub.group(1)
            continue

        m_main = RE_MAIN.match(text)
        if m_main:
            code, desc = m_main.group(1), m_main.group(2).strip()
            records.append({code: desc})
    return records


# Streamlit UI
########

st.set_page_config(page_title="Besidetech Excel Extractor", page_icon="📊")
st.title("Besidetech - Estrazione criteri XLS")

uploaded_file = st.file_uploader("📤 Carica file Excel", ["xls", "xlsx"])
if uploaded_file is None:
    st.info("Carica un file per continuare.")
    st.stop()


# Carica workbook e mostra fogli
# ---------------------------------------------------------------------------
try:
    wb = _workbook_from_upload(uploaded_file)
except Exception as exc:
    st.error(f"Errore apertura file: {exc}")
    st.stop()

sheet_names = wb.sheetnames
st.subheader("📄 Pagine e fogli rilevate:")
for i, name in enumerate(sheet_names, 1):
    st.write(f"**{i}. {name}**")

sheet_name = st.selectbox("Seleziona foglio per nome", sheet_names, index=0)
sheet_idx = st.number_input("…oppure per numero (1 = primo)", 1, len(sheet_names), 1, 1, format="%d")
if sheet_idx != 1:
    sheet_name = sheet_names[int(sheet_idx) - 1]

sheet = wb[sheet_name]

col_letter = st.text_input("Lettera colonna dati", value="A", max_chars=3).strip()
mode = st.radio("Intervallo di lettura", ["Tutta la colonna", "Intervallo di righe"], 0)

if mode == "Intervallo di righe":
    row_start = st.number_input("Riga inizio", 1, sheet.max_row, 1)
    row_end = st.number_input("Riga fine", int(row_start), sheet.max_row, sheet.max_row)
else:
    row_start, row_end = 1, sheet.max_row


# Chiave di configurazione per capire se dobbiamo ricalcolare
# ---------------------------------------------------------------------------
config_str = f"{uploaded_file.name}|{sheet_name}|{col_letter}|{row_start}-{row_end}"
config_hash = hashlib.sha1(config_str.encode()).hexdigest()

if st.session_state.get("config_hash") != config_hash:
    st.session_state.pop("records", None)
    st.session_state.pop("codes", None)
    st.session_state.pop("selected_codes", None)
    st.session_state["config_hash"] = config_hash

# Bottone analisi
# ---------------------------------------------------------------------------
if st.button("Analizza / Aggiorna") or st.session_state.get("records") is None:
    with st.spinner("Analisi in corso…"):
        records = _parse_records(sheet, col_letter, int(row_start), int(row_end))
        st.session_state["records"] = records
        st.session_state["codes"] = [next(iter(d)) for d in records]
        st.session_state["selected_codes"] = st.session_state["codes"]  # default all selected


# Se ci sono dei  records, mostriamo multiselect e download
# ---------------------------------------------------------------------------
records = st.session_state.get("records")
if records:
    st.success(f"Trovati {len(records)} codici nel foglio '{sheet_name}'.")

    selected = st.multiselect(
        "Seleziona i codici da esportare",
        options=st.session_state["codes"],
        default=st.session_state.get("selected_codes", st.session_state["codes"]),
        key="selected_codes",
    )

    filtered_records = [rec for rec in records if next(iter(rec)) in selected]

    if filtered_records:
        json_obj = {"estrazione": filtered_records}
        json_text = json.dumps(json_obj, ensure_ascii=False, indent=2)
        st.subheader("📤 Anteprima JSON")
        st.json(json_obj)
        st.download_button("💾 Scarica JSON", json_text, "estrazione.json", "application/json")
    else:
        st.warning("Nessun codice selezionato.")
